# Entire rewrite of server.py but using WSGI as the back
# Perfect for running in cPanel as I am trying to host this
# on VinaHost, which only supports WSGI and not HTTPD.

from sys import argv, stdout, stderr
import os
import mimetypes
import sqlite3
from urllib.parse import parse_qs, urlparse
import json
import datetime
import openpyxl.styles
from search import search_books, normalize_unicode, all_books
from hashlib import sha256
import base64
import random
import string
import sys
import subprocess
from multiprocessing import Process
import time
import openpyxl

db_name = "data"

db = sqlite3.connect(f"{db_name}.db")
cursor = db.cursor()

DEFAULT_FILE_PATH = './index.html'
LOG_FILE = "server_output.log"
STATIC_DIR = './src'

sys.path.insert(0, os.path.dirname(__file__))

def get_code_data(code: str):
    print("Processing code: " + code)
    try:
        cursor.execute('SELECT * FROM ids WHERE id=?', (code,))
        result = cursor.fetchone()
        cursor.execute('SELECT * FROM books WHERE id=?', (result[1],))
        result2 = cursor.fetchone()
    except Exception as e:
        result = ('null', 'null', 'null', 'null', 'null')
        result2 = ('null', 'null', 'null', 'null', 'null', 'null')
        print("Error in get_code_data() result1/2: "+str(e))

    try:
        cursor.execute('SELECT * FROM borrows WHERE id=?', (code,))
        result3 = cursor.fetchone()
        result3[1]
    except Exception as e:
        result3 = ('0', '0', 0, 0)
        print("Error in get_code_data() result3: "+str(e))

    try:
        cursor.execute('SELECT name FROM users WHERE username=?', (result3[1],))
        borrowername = cursor.fetchone()[0]
    except Exception as e:
        borrowername = result3[1]

    json_data = {
        "id": result[0],
        "type": result[1],
        "episode": result[2],
        "donator": result[3],
        "donation_day": result[4],
        "title": result2[1],
        "author": result2[2],
        "year_published": result2[3],
        "description": result2[4],
        "use": result2[5],
        "borrower": borrowername,
        "borrow_date": result3[2],
        "borrow_expire": result3[3]
    }
    return json.dumps(json_data, ensure_ascii=False).encode('utf-8')

def get_search_data(idstr: str):
    # Perform the search with the updated function
    print("Search:", idstr)
    valid_ids, detailed_results = search_books(cursor, idstr, 40)

    # Initialize lists to store the results
    bktypes = []
    ranks = []
    titles = []
    episodes = []

    # Iterate over the detailed results
    for valid_id, book_id, rank in detailed_results:
        # Append the data to respective lists
        bktypes.append(book_id)  # Original book ID (type)
        ranks.append(rank)       # Confidence rank

        # Fetch the title for the current book_id
        cursor.execute("SELECT title FROM books WHERE id=?", (book_id,))
        title_result = cursor.fetchone()
        titles.append(title_result[0] if title_result else "Unknown")  # Handle missing titles

        # Fetch the episode information for the valid_id
        cursor.execute("SELECT ep FROM ids WHERE id=?", (valid_id,))
        episode_result = cursor.fetchone()
        episodes.append(episode_result[0] if episode_result else "N/A")  # Handle missing episodes

    # Build the JSON data
    json_data = {
        "valid_ids": valid_ids,
        "bktypes": bktypes,
        "ranks": ranks,
        "titles": titles,
        "episodes": episodes
    }

    # Return the JSON data as a UTF-8 encoded string
    return json.dumps(json_data, ensure_ascii=False).encode('utf-8')

def get_user_data(username: str):
    cursor.execute("SELECT * FROM users WHERE username=?", (username,))
    userdata = cursor.fetchone() # ("nautilus", "123", "Nautilus4K", 1732372092, None)
    # print(userdata)
    if (userdata == None):
        userdata = [None, None, None, None, None]
    else: userdata = list(userdata)
    if (userdata[2] == None): userdata[2] = userdata[0]

    json_data = {
        "username": userdata[0],
        # "password": userdata[1],  # Dangerous. Could be used to some degree of advantages
        "name": userdata[2],
        "date": userdata[3],
        # "phone": userdata[4]  # Uneeded. Probably just wasted some more bandwidths with this on.
                                # Also has to limit the amount of personal data could be accessed
                                # via API calls
    }

    return json.dumps(json_data, ensure_ascii=False).encode('utf-8')

def check_login_credentials(username: str, passwd: str):
    hashedpass = sha256(passwd.encode("utf-8")).hexdigest()
    cursor.execute("SELECT * FROM users WHERE username=?", (username,))
    userdata = cursor.fetchone() # ("nautilus", "123", "Nautilus4K", 1732372092, None)
    corr = True

    # If no user is found to be matched
    if (userdata == None):
        corr = False
    elif (userdata[1] != hashedpass):
        corr = False

    json_data = {
        "username": username,
        "correct": corr
    }

    return json.dumps(json_data, ensure_ascii=False).encode('utf-8')

def modify_account(username: str, passwd: str, newpass: str):
    # print("Hashing...")
    hashedpass = sha256(passwd.encode("utf-8")).hexdigest()
    hashednewpass = sha256(newpass.encode("utf-8")).hexdigest()

    # print("Getting information...")
    cursor.execute("SELECT * FROM users WHERE username=?", (username,))
    userdata = cursor.fetchone()

    # print("Checking integrity...")
    error = False
    if (userdata == None): error = True
    elif (userdata[1] != hashedpass): error = True
    else:
        # print("Updating...")
        cursor.execute("UPDATE users SET password = ? WHERE username = ?", (hashednewpass, username))
        db.commit()

    # print("Sending back...")
    json_data = {
        "username": username,
        "error": error,
    }

    return json.dumps(json_data, ensure_ascii=False).encode('utf-8')

def generate_token():
    # Characters to choose from: uppercase, lowercase letters, and digits
    characters = string.ascii_letters + string.digits
    # Create 5 groups of 5 characters joined by dashes
    token = "-".join(
        "".join(random.choices(characters, k=5)) for _ in range(5)
    )
    return token

def get_facial_result(token: str):
    json_data = {}
    currpath = os.getcwd()+"/faceserver_workspaces/result/"
    if os.path.exists(currpath+token+".json"):
        f = open(currpath+token+".json")
        facialresult = json.loads(f.read())

        # If result exists
        json_data = {
            "status": "SUCCESSFUL",
            "error": facialresult["error"],
            "result": facialresult["result"]
        }
    else:
        json_data = {
            "status": "WAITING",
            "error": None,
            "result": None
        }

    # print(currpath+token+".json")
    return json.dumps(json_data, ensure_ascii=False).encode('utf-8')

def borrow_book(code: str, username: str):
    json_data = {}
    print("Now borrowing book: "+code)
    current_time = int(time.time())  # Get current Unix timestamp
    expire_time = current_time + 7 * 24 * 60 * 60  # Borrow period of 7 days from now
    data = (code, username, current_time, expire_time)

    # Insert data into the table
    try:
        cursor.execute('''
            INSERT OR IGNORE INTO  borrows (id, current_borrower, borrow_day, borrow_expire)
            VALUES (?, ?, ?, ?)
        ''', data)
        json_data = {
            "success": True
        }
    except Exception as e:
        print("Error inserting data into database: ", e)
        json_data = {
            "success": False
        }

    db.commit()
    return json.dumps(json_data, ensure_ascii=False).encode('utf-8')    

def run_facial_reg():
    # Run facial recognition server. Which will be used to get facial recoginition functionality inside of log in page.
    print("Execution facial recoginition script...")
    subprocess.run(['python', 'face_server.py'])

global row
def save_to_excel(ip_address: str, username: str, name: str, time_act: str, action: str):
    row = 3
    # Define a thin border
    thin_border = openpyxl.styles.Border(
        left=openpyxl.styles.Side(style='thin'),
        right=openpyxl.styles.Side(style='thin'),
        top=openpyxl.styles.Side(style='thin'),
        bottom=openpyxl.styles.Side(style='thin')
    )
    if os.path.exists("misc/written.xlsx"):
        workbook = openpyxl.load_workbook("misc/written.xlsx")
        sheet = workbook.active
        workbook.save("misc/written.xlsx")

        if row == 3:
            # Continually find until found a blank row
            while True:
                cell_value = sheet[f"A{row}"].value  # Get the value of the current cell in column A
                if cell_value is None:  # Check if the cell is empty
                    break  # Stop the loop if the cell is blank
                
                row += 1  # Move to the next row (A4, A5, etc.)
    else:

        # Create a new workbook and add a worksheet
        if not os.path.exists("misc"): os.makedirs("misc")
        workbook = openpyxl.Workbook()  # Create a new workbook
        sheet = workbook.active  # Get the active sheet
        sheet.title = "LogSheet"  # Rename the sheet (optional)

        # Preparation
        sheet["A1"] = "Nhật ký kết nối thư viện"
        sheet["A1"].font = openpyxl.styles.Font(name="Times New Roman", size=18, bold=True)

        # Now title rows, which stands for itself as one of the best rows in the entire workbook
        sheet["A2"] = "IP"
        sheet["A2"].font = openpyxl.styles.Font(name="Times New Roman", size=11, bold=True, color="000000")
        sheet["A2"].alignment = openpyxl.styles.Alignment(wrap_text=True, horizontal="center", vertical="center")
        sheet["A2"].fill = openpyxl.styles.PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        sheet["A2"].border = thin_border

        sheet["B2"] = "Định danh"
        sheet["B2"].font = openpyxl.styles.Font(name="Times New Roman", size=11, bold=True, color="000000")
        sheet["B2"].alignment = openpyxl.styles.Alignment(wrap_text=True, horizontal="center", vertical="center")
        sheet["B2"].fill = openpyxl.styles.PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        sheet["B2"].border = thin_border

        sheet["C2"] = "Tên người dùng"
        sheet["C2"].font = openpyxl.styles.Font(name="Times New Roman", size=11, bold=True, color="000000")
        sheet["C2"].alignment = openpyxl.styles.Alignment(wrap_text=True, horizontal="center", vertical="center")
        sheet["C2"].fill = openpyxl.styles.PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        sheet["C2"].border = thin_border

        sheet["D2"] = "Thời gian"
        sheet["D2"].font = openpyxl.styles.Font(name="Times New Roman", size=11, bold=True, color="000000")
        sheet["D2"].alignment = openpyxl.styles.Alignment(wrap_text=True, horizontal="center", vertical="center")
        sheet["D2"].fill = openpyxl.styles.PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        sheet["D2"].border = thin_border

        sheet["E2"] = "Action"
        sheet["E2"].font = openpyxl.styles.Font(name="Times New Roman", size=11, bold=True, color="000000")
        sheet["E2"].alignment = openpyxl.styles.Alignment(wrap_text=True, horizontal="center", vertical="center")
        sheet["E2"].fill = openpyxl.styles.PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        sheet["E2"].border = thin_border

        sheet.column_dimensions['A'].width = 16
        sheet.column_dimensions['B'].width = 12
        sheet.column_dimensions['C'].width = 25
        sheet.column_dimensions['D'].width = 20
        sheet.column_dimensions['E'].width = 70

        workbook.save("misc/written.xlsx")

    # Now we actually add informations into the EXCEL file.

    sheet[f"A{row}"] = ip_address
    sheet[f"A{row}"].font = openpyxl.styles.Font(name="Times New Roman", size=11, bold=True, color="000000")
    sheet[f"A{row}"].border = thin_border
    sheet[f"A{row}"].alignment = openpyxl.styles.Alignment(wrap_text=True, horizontal="center", vertical="center")
    sheet[f"A{row}"].fill = openpyxl.styles.PatternFill(start_color="88FFAA", end_color="88FFAA", fill_type="solid")

    sheet[f"B{row}"] = username
    sheet[f"B{row}"].font = openpyxl.styles.Font(name="Times New Roman", size=11, bold=False, color="000000")
    sheet[f"B{row}"].border = thin_border
    sheet[f"B{row}"].alignment = openpyxl.styles.Alignment(wrap_text=True, horizontal="center", vertical="center")

    sheet[f"C{row}"] = name
    sheet[f"C{row}"].font = openpyxl.styles.Font(name="Times New Roman", size=11, bold=False, color="000000")
    sheet[f"C{row}"].border = thin_border
    sheet[f"C{row}"].alignment = openpyxl.styles.Alignment(wrap_text=True, horizontal="center", vertical="center")

    if username == "null": 
        sheet[f"B{row}"].fill = openpyxl.styles.PatternFill(start_color="FF4242", end_color="FF4242", fill_type="solid")
        sheet[f"C{row}"].fill = openpyxl.styles.PatternFill(start_color="FF4242", end_color="FF4242", fill_type="solid")

    sheet[f"D{row}"] = time_act
    sheet[f"D{row}"].font = openpyxl.styles.Font(name="Times New Roman", size=11, bold=False, color="000000")
    sheet[f"D{row}"].border = thin_border
    sheet[f"D{row}"].alignment = openpyxl.styles.Alignment(wrap_text=True, horizontal="center", vertical="center")

    sheet[f"E{row}"] = action
    sheet[f"E{row}"].font = openpyxl.styles.Font(name="Times New Roman", size=11, bold=False, color="000000")
    sheet[f"E{row}"].border = thin_border
    sheet[f"E{row}"].alignment = openpyxl.styles.Alignment(wrap_text=True, horizontal="center", vertical="center")

    workbook.save("misc/written.xlsx")

    row += 1 # Saving this row for caching.

def log_activity(ip_address: str ,username: str, action: str):
    # This function is used to log the activity of the user.
    # It will save the activity into the EXCEL file.
    # The EXCEL file is located in the misc folder.
    json_data = {}
    try:
        cursor.execute('SELECT name FROM users WHERE username=?', (username,))
        name = cursor.fetchone()[0]
    except Exception as e:
        name = username
    timenow = datetime.datetime.now()
    timeact = timenow.strftime("%H:%M:%S %d-%m-%Y")
    if username != "admin": save_to_excel(ip_address, username, name, timeact, action)
    json_data = {
        "success": True,
    }
    return json.dumps(json_data, ensure_ascii=False).encode('utf-8')  

def get_stats():
    # Getting the statistics of the website
    json_data = {}

    try:
        # On book types
        cursor.execute("SELECT COUNT(*) FROM books WHERE use='Sách Tham Khảo'")
        typecount_thamkhao = cursor.fetchone()[0]
        cursor.execute("SELECT COUNT(*) FROM books WHERE use='Sách Truyện'")
        typecount_truyen = cursor.fetchone()[0]
        cursor.execute("SELECT COUNT(*) FROM books WHERE use='Sách Giáo Khoa'")
        typecount_giaokhoa = cursor.fetchone()[0]
        cursor.execute("SELECT COUNT(*) FROM books WHERE use='Sách Bài Tập'")
        typecount_baitap = cursor.fetchone()[0]
        json_data = {
            "success": True,
            "thamkhao": typecount_thamkhao,
            "truyen": typecount_truyen,
            "giaokhoa": typecount_giaokhoa,
            "baitap": typecount_baitap
        }
    except Exception as e:
        json_data = {
            "success": False,
            "error": str(e)
        }

    return json.dumps(json_data, ensure_ascii=False).encode('utf-8')

def id_generate(size, allowed_chars : str):
    return ''.join(random.choice(allowed_chars) for x in range(size)) 

def refresh_db(db):
    db.close()
    newdb = sqlite3.connect(f"{db_name}.db")
    return newdb

# add_book(headers_dict["ID"], headers_dict["Title"], headers_dict["Author"], headers_dict["Year"], headers_dict["Description"], headers_dict["Use"])
def add_book(idx: str, title: str, author: str, year: int, description: str, use: str, episode: int, body: str):

    json_data = {}
    try:
        # Get a clone of books and table for easy access
        cursor.execute("SELECT * FROM books")
        books = list(cursor.fetchall())

        # Check if the book already exists
        if idx in books:
            json_data = {
                "success": False,
                "error": "Book already exists"
            }
            return json.dumps(json_data, ensure_ascii=False).encode('utf-8')
        else:
            # Insert the book into the database
            data = (idx, title, author, year, description, use)

            # Insert data into the table
            cursor.execute('''
                INSERT OR IGNORE INTO books (id, title, author, year, description, use)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', data)
            db.commit()
            json_data = {
                "success": True
            }

            body_data = base64.b64decode(body)
            # Insert sample data using current time for `borrow_day` and future time for `borrow_expire`
            current_time = int(time.time())  # Get current Unix timestamp

            data = (id_generate(11, string.ascii_letters+string.digits), idx, episode, "Nhà Trường", current_time)

            # Insert data into the table
            cursor.execute('''
                INSERT OR IGNORE INTO ids (id, type, ep, donor, donateday)
                VALUES (?, ?, ?, ?, ?)
            ''', data)

            db.commit()
            
            with open("book_thumbnails/" + idx + "_" + str(episode) + ".jpg", "wb") as file:
                file.write(body_data)

            db = refresh_db(db)
        
    except Exception as e:
        json_data = {
            "success": False,
            "error": str(e)
        }

    return json.dumps(json_data, ensure_ascii=False).encode('utf-8')

def add_id(idx: str, episode: int, body: str):
    json_data = {}  # Default empty json_data
    print(f"Add book with id {idx}, episode {episode}")
    try:
        if body == "":
            # No body detected, check if the episode already exists in the database
            cursor.execute("SELECT * FROM ids WHERE ep = ?", (episode,))
            if not cursor.fetchone():
                # Episode does not exist
                json_data = {
                    "success": False,
                    "error": f"Không có ảnh cho tập {episode}"
                }
            else:
                # Episode exists, insert sample data
                current_time = int(time.time())  # Get current Unix timestamp

                data = (id_generate(11, string.ascii_letters + string.digits), idx, episode, "Nhà Trường", current_time)

                # Insert data into the table
                cursor.execute('''
                    INSERT OR IGNORE INTO ids (id, type, ep, donor, donateday)
                    VALUES (?, ?, ?, ?, ?)
                ''', data)
                
                json_data = {
                    "success": True,
                }
                db.commit()  # Commit to save the changes

        else:
            # Body detected, insert data into the table like normal
            # Episode exists, insert sample data
            current_time = int(time.time())  # Get current Unix timestamp

            data = (id_generate(11, string.ascii_letters + string.digits), idx, episode, "Nhà Trường", current_time)

            # Insert data into the table
            cursor.execute('''
                INSERT OR IGNORE INTO ids (id, type, ep, donor, donateday)
                VALUES (?, ?, ?, ?, ?)
            ''', data)
            
            json_data = {
                "success": True,
            }
            db.commit()  # Commit to save the changes

            body_data = base64.b64decode(body)
            with open("book_thumbnails/" + idx + "_" + str(episode) + ".jpg", "wb") as file:
                file.write(body_data)
    except Exception as e:
        json_data = {
            "success": False,
            "error": str(e)
        }

    return json.dumps(json_data, ensure_ascii=False).encode('utf-8')

def delete_borrow(idx: str):
    # Deleting borrows from DB
    try:
        cursor.execute("DELETE FROM borrows WHERE id=?", (idx,))
        db.commit()
        json_data = {
            "success": True,
        }
    except Exception as e:
        json_data = {
            "success": False,
            "error": str(e)
        }

    return json.dumps(json_data, ensure_ascii=False).encode('utf-8')

def application(environ, start_response):
    
    # Get the HTTP method (GET or POST)
    method = environ.get('REQUEST_METHOD', 'GET')

    # Get the path from the URL
    path = environ.get('PATH_INFO', '/')

    # Get the user's IP address
    ip_address = environ.get('REMOTE_ADDR', 'Unknown')

    # Prepare the response headers
    headers = [('Content-Type', 'text/plain')]
    
    # Log the details to the log file
    LOGFILE = "serverLog.log"
    with open(LOGFILE, "a") as file:
        file.write(f"[{ip_address}] --- {method} {path}\n")
    
    # Helper function to extract headers as a string
    def get_headers(environ):
        headers_str = ''
        for key, value in environ.items():
            if key.startswith('HTTP_'):
                header_name = key[5:].replace('_', '-').title()  # Convert to standard header format
                headers_str += f'{header_name}: {value}\n'
        return headers_str
    
    def get_headers_dict(environ):
        headers = {}
        for key, value in environ.items():
            if key.startswith('HTTP_'):  # HTTP headers are prefixed with 'HTTP_'
                header_name = key[5:].replace('_', '-').title()  # Convert to standard header format
                headers[header_name] = value
        return headers

    # Handle GET request
    if method == 'GET':
        # Get headers as a string
        try:
            headers_dict = get_headers_dict(environ)
            headers_str = get_headers(environ)
        except:
            print("An error occured in headers variable creation")
            headers_dict = {}
            headers_str = ""
        
        # # Query string parameters
        query_params = environ.get('QUERY_STRING', '')
        body_str = f'Query string: {query_params}\n' if query_params else ''
        
        # start_response('200 OK', headers)
        # message = 'This is a GET request.\n'
        # response = message + headers_str + body_str

        if path == "/":
            headers.append(('Location', '/about.html'))
            start_response('302 Found', headers)
            return [b'Document moved to /about.html']
        elif path == "/get":
            # GET requests for informations
            if "Request" in headers_dict:
                # Answers code request
                start_response('200 OK', headers)
                return [b''+get_code_data(headers_dict["Request"])]
            elif "Search" in headers_dict:
                # Answers search request
                start_response('200 OK', headers)
                return [b''+get_search_data(headers_dict["Search"])]
            elif "Username" in headers_dict:
                # Answers username request
                start_response('200 OK', headers)
                return [b''+get_user_data(headers_dict["Username"])]
            elif "User-Agent" in headers_dict and not "curl" in headers_dict["User-Agent"]:
                # Redirect back into website if user is trying to connect through a browser
                headers.append(('Location', '/'))
                start_response('302 Found', headers)
                return [b'Please do not use a browser to connect to this feature of the website as it is only used for highly technically advanced softwares.\n'+
                        b'Made by Nguyen Van Quang Vinh in his 8th grade with love']
            else:
                # If no valid request, returns bad request
                headers.append(('Location', '/about.html'))
                start_response('400 Bad Request', headers)
                return [b'No valid headers found in this request.\nHeaders: \n-----------------------\n'+headers_str.encode()]
        
        elif path == "/logincheck":
            # Check if user is trying to log in with the correct credentials
            if "Username" in headers_dict and "Password" in headers_dict:
                # If all data exists, then we check and returns the result
                start_response('200 OK', headers)
                return [b''+check_login_credentials(headers_dict["Username"], headers_dict["Password"])]
            else:
                # If not all data exists, then we return bad request
                # headers.append(('Location', '/about.html'))
                start_response('400 Bad Request', headers)
                return [b'{"error": "Missing headers for connection"}']
            
        elif path == "/modifyaccounts":
            # Modify user accounts with this GET request. Definitely not a hoax. I promise.
            if "Username" in headers_dict and "Password" in headers_dict and "Newpass" in headers_dict:
                # If all data exists, then we check and returns the result
                start_response('200 OK', headers)
                return [b''+modify_account(headers_dict["Username"], headers_dict["Password"], headers_dict["Newpass"])]
            else:
                # If not all data exists, then we return bad request
                start_response('400 Bad Request', headers)
                return [b'{"error": "Missing headers for connection"}']
            
        elif path == "/getfacialresult":
            # Get facial results from faceserver_workspaces/result after processing from face_server.
            if 'Token' in headers_dict:
                # If token exists, then we check and returns the result of our very deep investigation
                start_response('200 OK', headers)
                return [b''+get_facial_result(headers_dict["Token"])]
            else:
                # If not all data exists, then we return bad request
                print("Bad request in /getfacialresult")
                start_response('400 Bad Request', headers)
                return [json.dumps({"error": "Missing headers for connection", "headers": headers_dict}).encode('utf-8')]
            
        elif path == "/borrowbook":
            # Borrow books from the library through API calls
            if 'Code' in headers_dict and 'Username' in headers_dict:
                # If code exists, then we mark the book as borrowed then returns if the process completed successfully or no.
                start_response('200 OK', headers)
                return [b''+borrow_book(headers_dict["Code"], headers_dict["Username"])]
            else:
                # If not all data exists, then we return bad request
                start_response('400 Bad Request', headers)
                return [json.dumps({"error": "Missing headers for connection", "headers": headers_dict}).encode('utf-8')]
            
        elif path == "/journal":
            # Logging user activity
            if 'Username' in headers_dict and 'Action' in headers_dict:
                # If all required params are present, then we log the activity
                start_response('200 OK', headers)
                return [b''+log_activity(ip_address, headers_dict["Username"], headers_dict["Action"])]
            
        elif path == "/getstats":
            start_response('200 OK', headers)
            return [b''+get_stats()]
        
        elif path == "/delborrow":
            # Delete borrowed book from the library through API calls
            if "Id" in headers_dict:
                start_response('200 OK', headers)
                return [b''+delete_borrow(headers_dict["Id"])]
            else:
                start_response('400 Bad Request', headers)
                return [json.dumps({"error": "Missing headers for connection", "headers": headers_dict}).encode('utf-8')]
        
        else:
            # This is for web access
            file_path = os.path.join(os.getcwd(), path.lstrip('/'))  # Get the file path
            if os.path.isfile(file_path):  # Check if the file exists
                # Get the file MIME type based on its extension
                mime_type, _ = mimetypes.guess_type(file_path)
                mime_type = mime_type or 'application/octet-stream'  # Default to binary if unknown

                # Set the appropriate Content-Type header
                headers[0] = ('Content-Type', mime_type)

                try:
                    # Check if the file is a text file (based on its MIME type)
                    if mime_type.startswith('text'):
                        # Automatically detect the file's encoding
                        with open(file_path, 'rb') as f:
                            raw_data = f.read()
                            detected_encoding = "utf-8"
                        
                        # Read the file content with the detected encoding
                        with open(file_path, 'r', encoding=detected_encoding) as f:
                            file_data = f.read()
                        
                        # Encode text content to bytes using UTF-8 for consistent delivery
                        file_data = file_data.encode('utf-8')
                        
                        # Update Content-Type to include charset for text files
                        headers[0] = ('Content-Type', f"{mime_type}; charset=utf-8")

                    else:
                        # For non-text files, read them in binary mode
                        with open(file_path, 'rb') as f:
                            file_data = f.read()
                    
                    # Set the Content-Length header
                    headers.append(('Content-Length', str(len(file_data))))

                    # Add Content-Disposition header to suggest inline display
                    headers.append(('Content-Disposition', f'inline; filename="{os.path.basename(file_path)}"'))

                    # Return the file content as the response
                    start_response('200 OK', headers)
                    return [file_data]

                except UnicodeDecodeError:
                    # If there is an issue with decoding, return a 500 error
                    start_response('500 Internal Server Error', headers)
                    return [b"Error decoding file content."]
                except IOError:
                    # If the file can't be read, return a 500 error
                    start_response('500 Internal Server Error', headers)
                    return [b"Error reading file."]

            else:
                # File not found, return 404
                start_response('404 Not Found', headers)
                return [b"File not found."]

    # Handle POST request
    elif method == 'POST':
        # Get headers as a string
        try:
            headers_dict = get_headers_dict(environ)
            headers_str = get_headers(environ)
        except:
            print("An error occured in headers variable creation")
            headers_dict = {}
            headers_str = ""
        if path == "/facial":
            # Handle facial recognition request
            try:
                # Get POST body data
                content_length = int(environ.get('CONTENT_LENGTH', 0))
                post_data = environ['wsgi.input'].read(content_length) if content_length > 0 else b''
                body_str = post_data.decode()
                image = base64.b64decode(body_str)
                newtoken = generate_token()
                # Save the decoded data as an image file
                with open(f'faceserver_workspaces/queue/{newtoken}.jpg', 'wb') as f:
                    f.write(image)
                print(f"Image received and saved with token: {newtoken}.")
                # Send success response
                response = {
                    'success': True,
                    'token': newtoken
                }
                start_response('200 OK', headers)
                return [b''+json.dumps(response).encode('utf-8')]
            except Exception as e:
                # If there is an issue with decoding, return an error
                response = {
                    'success': False,
                    'error': str(e)
                }

        elif path == "/addbook":
            if 'Id' in headers_dict and 'Title' in headers_dict and 'Author' in headers_dict and 'Year' in headers_dict and 'Description' in headers_dict and 'Use' in headers_dict and 'Episode' in headers_dict:
                # Read the content length
                content_length = int(environ.get('CONTENT_LENGTH', 0))
                
                # Read the request body as a string
                body_str = environ['wsgi.input'].read(content_length).decode('utf-8')
                start_response('200 OK', headers)
                return [b''+add_book(headers_dict["Id"], headers_dict["Title"], headers_dict["Author"], headers_dict["Year"], headers_dict["Description"], headers_dict["Use"], headers_dict["Episode"], body_str)]
            elif 'Id' in headers_dict and 'Episode' in headers_dict and headers_dict['Id'] != "" and headers_dict['Episode'] != "":
                # Read the content length
                content_length = int(environ.get('CONTENT_LENGTH', 0))
                start_response('200 OK', headers)
                if content_length == 0:
                    body_str = ''
                    print("No body detected on connection")
                else:
                    body_str = environ['wsgi.input'].read(content_length).decode('utf-8')
                print(headers_dict["Id"], headers_dict["Episode"])
                return [b''+add_id(headers_dict["Id"], headers_dict["Episode"], body_str)]
            else:
                start_response('400 Bad Request', headers)
                print(headers_dict)
                return [json.dumps({"error": "Missing headers for connection", "headers": headers_dict}).encode('utf-8')]
            
        else:
            # Get headers as a string
            headers_str = get_headers(environ)
            
            # Get POST body data
            content_length = int(environ.get('CONTENT_LENGTH', 0))
            post_data = environ['wsgi.input'].read(content_length) if content_length > 0 else b''
            body_str = post_data.decode()

            # Parse POST data (application/x-www-form-urlencoded)
            post_params = parse_qs(body_str)
            
            start_response('200 OK', headers)
            message = 'No POST request could be sent to this part of the website.\n'
            response = message + headers_str + f'POST data: \n---------------------\n{body_str}\n'

            # Optionally, show parsed parameters as well
            if post_params:
                response += '\nParsed POST data:\n---------------------\n'
                for key, value in post_params.items():
                    response += f'{key}: {value}\n'

    else:
        # Handle other HTTP methods (e.g., PUT, DELETE)
        start_response('405 Method Not Allowed', headers)
        response = 'Method Not Allowed\n'

    # Return the response as a byte-encoded list
    return [response.encode('utf-8')]

# if __name__ == "__main__":
    # Start facial recoginition server.
# print("Executing facial recoginition server...")
# face_proc = Process(target=run_facial_reg)
# face_proc.start()